import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from datetime import datetime


def generate_daily_report(source_file, report_file, gross_wb, target_date_input):

    # === CONFIGURATION ===

    # Usage in your script:
    target_date = pd.to_datetime(target_date_input)
    date_str = target_date.strftime("%dth %B")  # Formats as "9th March"

    # === Load Source Data ===
    source_file = "03 March  25 Gross Gen.xlsx"
    summary_df = pd.read_excel(source_file, sheet_name="Summary", header=None)

    # Make a full copy for lookups
    full_df = summary_df.copy()

    # Extract data section and clean it
    data_df = summary_df.iloc[9:].copy()

    print("\n=== Checking if 'Total MWH' is in raw summary_df ===")
    matches = summary_df.apply(lambda row: row.astype(str).str.contains("Total MWH", case=False, na=False).any(), axis=1)
    total_mwh_rows = summary_df[matches]
    print(total_mwh_rows)


    # Reset index
    data_df = data_df.reset_index(drop=True)

    # Rename columns manually
    data_df.columns = [
        "DATE", "DG SET .1", "DG SET .2", "DG SET .3", "DG SET .4", 
        "DG SET .5", "STG", "DAILY TOTAL MWH", "PLANT GROSS", "ENG GROSS", "STG2"
    ]

    # Improved date parsing - handle multiple formats
    data_df["DATE_PARSED"] = pd.to_datetime(
        data_df["DATE"], 
        errors='coerce',
        dayfirst=True  # Important for European date formats
    ).dt.normalize()  # Remove time component

    # Debug: Show available dates
    print("Available dates in source data:")
    print(data_df["DATE_PARSED"].dropna().dt.strftime("%Y-%m-%d").unique())


    # Find the MWH value for the date
    daily_row = data_df[data_df["DATE_PARSED"] == target_date]
    mwh_value = daily_row["DAILY TOTAL MWH"].values[0] if not daily_row.empty else None

    def extract_monthly_mwh(summary_df, total_label="Total MWH"):
        print("\n=== Debugging extract_monthly_mwh() ===")
        
        # Search for row that contains the label (excluding header rows)
        mask = summary_df.apply(lambda row: row.astype(str).str.strip().str.lower().str.contains(total_label.lower()).any(), axis=1)
        # Skip first 10 rows which likely contain headers
        mask[:10] = False
        total_row = summary_df[mask]
        
        if total_row.empty:
            print(f"No row found containing '{total_label}'")
            return None
        
        print(f"Found row with label '{total_label}':")
        print(total_row)
        
        # Find the "DAILY TOTAL MWH" column index from header row (row 7)
        daily_total_col = None
        header_row = summary_df.iloc[7]  # Headers are in row 8 (0-based index 7)
        for col_idx, val in enumerate(header_row):
            if str(val).strip().upper() == "DAILY TOTAL MWH":
                daily_total_col = col_idx
                print(f"Found 'DAILY TOTAL MWH' at column index {col_idx}")
                break
        
        if daily_total_col is not None:
            # Get the value from Total MWH row (not header row)
            val = total_row.iloc[0, daily_total_col]
            print(f"Value in same column of Total MWH row: {val} (type: {type(val)})")
            try:
                float_val = float(val)
                print(f"Monthly MWH from '{total_label}' row: {float_val}")
                return float_val
            except (ValueError, TypeError):
                print(f"Could not convert value '{val}' to float")
        
        # Fallback: Check all columns for numeric values in the actual Total MWH row
        print("\nChecking all columns for numeric values:")
        for col in reversed(total_row.columns):
            val = total_row.iloc[0][col]
            print(f"Column {col}: {val} (type: {type(val)})")
            try:
                if pd.notna(val):
                    float_val = float(val)
                    print(f"Found numeric value: {float_val}")
                    return float_val
            except (ValueError, TypeError):
                continue
        
        print(f"No numeric value found in row labeled '{total_label}'")
        return None

    # === Write to Daily Report ===
    if mwh_value is not None:
        report_file = "Daily production report March 2025.xlsx"
        wb = load_workbook(report_file)
        
        # Check if a sheet with the given date_str exists
        if date_str in wb.sheetnames:
            # If the sheet exists, update the value in cell B8
            ws = wb[date_str]
            ws["B8"] = mwh_value
            wb.save("Daily production report March 2025.xlsx")
            print(f"Inserted {mwh_value} into sheet '{date_str}' cell B8.")
        
        # If the sheet doesn't exist, create a new sheet with the date_str name
        else:
            # Find an existing sheet to duplicate (use the first sheet here as an example)
            sheet_to_duplicate = wb.worksheets[2]  # You can adjust this if you want a specific sheet
            
            # Create a new sheet by copying the content of the original sheet
            new_sheet = wb.copy_worksheet(sheet_to_duplicate)

            
            
            # Rename the new sheet to the current date_str
            new_sheet.title = date_str

            for row in sheet_to_duplicate.iter_rows():
                for cell in row:
                # Skip MergedCells that are not the actual top-left anchor
                    if isinstance(cell, MergedCell):
                        continue

                    new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                                            
                    # Copy styles
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

                
                    # Copy only formulas or static labels (no user-filled values)
                    if cell.data_type == 'f':
                        new_cell.value = f"={cell.value}"
                    elif isinstance(cell.value, str) and cell.value.strip() != "":
                        new_cell.value = cell.value  # Copy headers/static text
                    else:
                        new_cell.value = None  # Clear user-entered numbers or blanks

            # Copy merged cell ranges
            for merged_range in sheet_to_duplicate.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_range))
            # Copy column widths
            for col in sheet_to_duplicate.column_dimensions:
                new_sheet.column_dimensions[col].width = sheet_to_duplicate.column_dimensions[col].width

            # Copy row heights
            for row_dim in sheet_to_duplicate.row_dimensions:
                new_sheet.row_dimensions[row_dim].height = sheet_to_duplicate.row_dimensions[row_dim].height

            #Rename cell B3 with correct name
            new_sheet["B3"].value = f"{date_str} 2025"


            
            # Assign the new sheet to `ws` so we can insert value below
            ws = new_sheet
            print(f"New sheet '{date_str}' has been created.")
            
    

        # Insert the value into B8, regardless of whether the sheet existed or was just created
        ws["B8"] = mwh_value

        # Insert the monthly total MWH into B9
        monthly_mwh = extract_monthly_mwh(summary_df)  # Call your earlier function here
        if monthly_mwh is not None:
            ws["B9"] = monthly_mwh
            print(f"Inserted monthly MWH value {monthly_mwh} into cell B9.")
        else:
            print("Monthly MWH value not found. Skipping B9 insertion.")

        # Step: Extract Annual MWH from Gross Gen Summary 2025
        try:
            gross_wb = load_workbook("Gross Gen. Summary 2025.xlsx", data_only=True)
            gross_ws = gross_wb.active  # Or specify by name if needed: gross_wb["SomeSheetName"]
            annual_mwh_value = float(gross_ws["H17"].value)
            print(f"Annual MWH value extracted from H17: {annual_mwh_value}")
        except Exception as e:
            annual_mwh_value = None
            print(f"Error extracting annual MWH value: {e}")

        # Step: Add to B10 in the Daily Report
        if annual_mwh_value is not None:
            ws["B10"] = annual_mwh_value
            print(f"Inserted annual MWH value {annual_mwh_value} into cell B10.")
        else:
            print("Annual MWH value not found. Skipping B10 insertion.")


       

        print(f"Inserted {mwh_value} into sheet '{date_str}' cell B8.")
    else:
        print("No data found for that date.")

    # === Final output ===
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Rewind the buffer
    return output
